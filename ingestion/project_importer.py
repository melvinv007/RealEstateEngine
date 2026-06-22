"""
project_importer.py

Imports projects from an Excel master sheet into the MongoDB projects collection.
"""

import argparse
import re
import sys
from pathlib import Path
from typing import Any
import csv

from openpyxl import load_workbook

sys.path.append(str(Path(__file__).resolve().parent.parent))

from core.config import PROJECTS_MASTER_EXCEL
from core.database import insert_project

EXPECTED_COLUMNS = [
    "Youtube",
    "ProjectName",
    "Developer",
    "AreaName",
    "PropertyType",
    "StartingPrice",
    "ImageLink",
    "Lat,Lng",
    "PDF",
    "RedSticker",
    "LifeStyle",
    "Handover",
    "Payment Plan",
    "Bedrooms",
]


def _normalize_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped or stripped.upper() == "N/A":
            return None
        return stripped
    return value


def _parse_property_types(value: Any) -> list[str]:
    value = _normalize_value(value)
    if value is None:
        return []
    if isinstance(value, str):
        parts = [p.strip() for p in value.split(",") if p.strip()]
    else:
        parts = [str(value).strip()]

    mapping = {
        "apartments": "apartment",
        "apartment": "apartment",
        "villas": "villa",
        "villa": "villa",
        "townhouses": "townhouse",
        "townhouse": "townhouse",
        "penthouses": "penthouse",
        "penthouse": "penthouse",
        "plots": "plot",
        "plot": "plot",
        "offices": "office",
        "office": "office",
        "studios": "studio",
        "studio": "studio",
        "warehouses": "warehouse",
        "warehouse": "warehouse",
    }

    normalized: list[str] = []
    for part in parts:
        key = part.lower()
        normalized.append(mapping.get(key, key))

    return sorted(set(normalized))


def _parse_bedrooms(value: Any) -> list[int]:
    value = _normalize_value(value)
    if value is None:
        return []

    parts = [p.strip() for p in str(value).split(",") if p.strip()]
    results: list[int] = []

    for part in parts:
        part_lower = part.lower()
        if "studio" in part_lower:
            results.append(0)
            continue

        match = re.search(r"(\d+)", part_lower)
        if match:
            results.append(int(match.group(1)))

    return sorted(set(results))


def _parse_starting_price(value: Any) -> float | None:
    value = _normalize_value(value)
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    cleaned = re.sub(r"[^0-9\.]", "", str(value))
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def _parse_lat_lng(value: Any) -> dict | None:
    value = _normalize_value(value)
    if value is None:
        return None

    if isinstance(value, str):
        parts = [p.strip() for p in value.split(",")]
        if len(parts) != 2:
            return None
        try:
            lat = float(parts[0])
            lng = float(parts[1])
        except ValueError:
            return None
    elif isinstance(value, (list, tuple)) and len(value) == 2:
        try:
            lat = float(value[0])
            lng = float(value[1])
        except ValueError:
            return None
    else:
        return None

    return {
        "type": "Point",
        "coordinates": [lng, lat],
    }


def _build_project_doc(row: dict[str, Any]) -> dict:
    def _get_col(name: str) -> Any:
        return row.get(name)

    youtube = _normalize_value(_get_col("Youtube"))
    link = _normalize_value(_get_col("Link"))
    project_name = _normalize_value(_get_col("ProjectName"))
    developer = _normalize_value(_get_col("Developer"))
    area_name = _normalize_value(_get_col("AreaName"))
    property_type_raw = _normalize_value(_get_col("PropertyType"))
    starting_price_raw = _normalize_value(_get_col("StartingPrice"))
    image_link = _normalize_value(_get_col("ImageLink"))
    lat_lng_raw = _normalize_value(_get_col("Lat,Lng"))
    language = _normalize_value(_get_col("Language"))
    tours = _normalize_value(_get_col("Tours"))
    pdf_link = _normalize_value(_get_col("PDF"))
    red_sticker = _normalize_value(_get_col("RedSticker"))
    green_sticker = _normalize_value(_get_col("Green Sticker"))
    new_hot = _normalize_value(_get_col("New/Hot"))
    free_sticker = _normalize_value(_get_col("Free Sticker"))
    tags = _normalize_value(_get_col("Tags"))
    lifestyle = _normalize_value(_get_col("LifeStyle"))
    handover = _normalize_value(_get_col("Handover"))
    payment_plan = _normalize_value(_get_col("Payment Plan"))
    bedrooms_raw = _normalize_value(_get_col("Bedrooms"))
    link_ru = _normalize_value(_get_col("LinkRu"))
    link_ar = _normalize_value(_get_col("LinkAr"))

    property_types = _parse_property_types(property_type_raw)
    bhk_options = _parse_bedrooms(bedrooms_raw)
    starting_price = _parse_starting_price(starting_price_raw)
    location_coords = _parse_lat_lng(lat_lng_raw)

    return {
        "Youtube": youtube,
        "Link": link,
        "ProjectName": project_name,
        "Developer": developer,
        "AreaName": area_name,
        "PropertyType": property_type_raw,
        "StartingPrice": starting_price_raw,
        "ImageLink": image_link,
        "Lat,Lng": lat_lng_raw,
        "Language": language,
        "Tours": tours,
        "PDF": pdf_link,
        "RedSticker": red_sticker,
        "GreenSticker": green_sticker,
        "NewHot": new_hot,
        "FreeSticker": free_sticker,
        "Tags": tags,
        "LifeStyle": lifestyle,
        "Handover": handover,
        "PaymentPlan": payment_plan,
        "Bedrooms": bedrooms_raw,
        "LinkRu": link_ru,
        "LinkAr": link_ar,
        "property_types": property_types,
        "bhk_options": bhk_options,
        "starting_price": starting_price,
        "location_coords": location_coords,
    }

def _iter_excel_rows(path: Path, start_row: int | None) -> list[dict[str, Any]]:
    """Legacy path: title row, header on row 2, data from row 3."""
    workbook = load_workbook(filename=path, data_only=True)
    sheet = workbook["Sheet 1"]
    header = [str(c.value).strip() if c.value is not None else "" for c in sheet[2]]

    rows: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=start_row or 3, values_only=True):
        if not row or not any(cell is not None and str(cell).strip() for cell in row):
            continue
        row_dict = {name: row[idx] if idx < len(row) else None for idx, name in enumerate(header) if name}
        rows.append(row_dict)
    return rows


def _iter_csv_rows(path: Path, start_row: int | None) -> list[dict[str, Any]]:
    """Plain CSV export: header is the first line, data starts right after."""
    csv.field_size_limit(sys.maxsize)
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader, start=1):
            if i < (start_row or 1):
                continue
            if not any(v is not None and str(v).strip() for v in row.values()):
                continue
            rows.append(row)
    return rows


def run_import(start_row: int | None = None, excel_path: str | None = None) -> dict:
    path = Path(excel_path or PROJECTS_MASTER_EXCEL)
    suffix = path.suffix.lower()

    if suffix == ".csv":
        rows = _iter_csv_rows(path, start_row)
    elif suffix in (".xlsx", ".xls", ".xlsm"):
        rows = _iter_excel_rows(path, start_row)
    else:
        raise ValueError(f"Unsupported projects master file type: {suffix or '(none)'}")

    inserted = 0
    skipped = 0

    for row_dict in rows:
        result = insert_project(_build_project_doc(row_dict))
        if result == "duplicate":
            skipped += 1
        else:
            inserted += 1

    print(f"[ProjectImporter] {inserted} inserted, {skipped} duplicates skipped")
    return {"inserted": inserted, "skipped": skipped}

def main() -> None:
    parser = argparse.ArgumentParser(description="Import projects from Excel or CSV")
    parser.add_argument("--start-row", type=int, default=None,
                         help="1-based start row. Defaults to 3 for .xlsx, 1 for .csv")
    parser.add_argument("--excel-path", type=str, default=None, help="Path to projects_master.csv or .xlsx")
    args = parser.parse_args()
    run_import(start_row=args.start_row, excel_path=args.excel_path)


if __name__ == "__main__":
    main()

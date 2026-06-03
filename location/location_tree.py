"""
location_tree.py

Build and cache an in-memory location tree with aliases and coordinates.
"""

import csv
import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any
import sys
csv.field_size_limit(sys.maxsize)

from openpyxl import load_workbook

EXCEL_PATH = "data/location_master_test.xlsx"
LOCATIONS_CSV = "data/locations.csv"
COORDINATES_CSV = "data/coordinates.csv"
TREE_CACHE = "cache/location_tree.json"
UNIQUE_NAMES_CACHE = "cache/unique_names.json"
FORCE_REBUILD = False

_NUM_SUFFIX_RE = re.compile(r"^(.*?)\s*(\d+)\s*$")


_TREE: dict[str, dict[str, Any]] | None = None
_ALIAS_MAP: dict[str, list[dict[str, Any]]] | None = None
_CANONICAL_MAP: dict[str, dict[str, Any]] | None = None
_LEVEL_SETS: dict[str, set[str]] | None = None


def normalize_key(value: str) -> str:
    """
    Lowercase -> strip -> remove dots -> collapse whitespace.
    """
    if value is None:
        return ""
    value = str(value).lower().strip()
    value = value.replace(".", "")
    value = re.sub(r"\s+", " ", value)
    return value


def split_phase(value: str) -> tuple[str, int | None]:
    """
    Split a normalized string into (base, phase_number).
    """
    key = normalize_key(value)
    if not key:
        return "", None

    match = _NUM_SUFFIX_RE.match(key)
    if match:
        base = match.group(1).strip()
        phase = int(match.group(2))
        if base:
            return base, phase

    concat = re.match(r"^([a-z][a-z\s]*)(\d+)$", key)
    if concat:
        base = concat.group(1).strip()
        phase = int(concat.group(2))
        if base:
            return base, phase

    return key, None


def _timestamp() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _load_tier3_set() -> set[str]:
    path = Path(UNIQUE_NAMES_CACHE)
    if not path.exists():
        return set()

    with path.open("r", encoding="utf-8") as file:
        data = json.load(file)

    tier3: set[str] = set()
    for entry in data.get("properties", []):
        if entry.get("tier3"):
            name = normalize_key(entry.get("canonical", ""))
            if name:
                tier3.add(name)
    return tier3


def _ensure_node(nodes: dict[str, dict[str, Any]], canonical: str, level: str | None, tier3_set: set[str]) -> dict[str, Any] | None:
    canonical = str(canonical or "").strip()
    if not canonical:
        return None

    norm = normalize_key(canonical)
    if not norm:
        return None

    node = nodes.get(norm)
    if not node:
        node = {
            "canonical": canonical,
            "level": level or "",
            "parent": None,
            "children": [],
            "aliases": [],
            "coords": None,
            "tier3": norm in tier3_set,
        }
        nodes[norm] = node
    else:
        if level and (not node.get("level") or node.get("level") == "unknown"):
            node["level"] = level

    if norm in tier3_set:
        node["tier3"] = True

    return node


def _add_child(parent: dict[str, Any], child: dict[str, Any]) -> None:
    if not parent or not child:
        return
    if not child.get("parent"):
        child["parent"] = parent.get("canonical")
    if child.get("canonical") not in parent.get("children", []):
        parent.setdefault("children", []).append(child.get("canonical"))


def _infer_parent_level(level: str) -> str | None:
    mapping = {
        "community": "city",
        "subcommunity": "community",
        "property": "subcommunity",
    }
    return mapping.get(level)


def _load_excel(nodes: dict[str, dict[str, Any]], tier3_set: set[str]) -> None:
    path = Path(EXCEL_PATH)
    if not path.exists():
        print(f"[{_timestamp()}] Warning: Excel not found: {EXCEL_PATH}")
        return

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        print(f"[{_timestamp()}] Warning: Excel sheet is empty.")
        return

    header_map: dict[str, int] = {}
    for idx, name in enumerate(header):
        key = normalize_key(name or "")
        if key:
            header_map[key] = idx

    def _col(name: str) -> int | None:
        key = normalize_key(name)
        return header_map.get(key)

    def _cell_str(val):
        return str(val).strip() if val is not None else ""

    col_city = _col("City")
    col_community = _col("Community")
    col_subcommunity = _col("Subcommunity")
    col_property = _col("Property")



    if col_city is None or col_community is None or col_subcommunity is None:
        print(f"[{_timestamp()}] Warning: Missing required columns in Excel; skipping hierarchy build.")
        return

    for row in rows:
        city = _cell_str(row[col_city]) if col_city is not None else ""
        community = _cell_str(row[col_community]) if col_community is not None else ""
        subcommunity = _cell_str(row[col_subcommunity]) if col_subcommunity is not None else ""
        property_ = _cell_str(row[col_property]) if col_property is not None else ""

        city_node = _ensure_node(nodes, city, "city", tier3_set) if city else None
        community_node = _ensure_node(nodes, community, "community", tier3_set) if community else None
        subcommunity_node = _ensure_node(nodes, subcommunity, "subcommunity", tier3_set) if subcommunity else None
        property_node = _ensure_node(nodes, property_, "property", tier3_set) if property_ else None

        if city_node and community_node:
            _add_child(city_node, community_node)
        if community_node and subcommunity_node:
            _add_child(community_node, subcommunity_node)
        if subcommunity_node and property_node:
            _add_child(subcommunity_node, property_node)
        elif community_node and property_node and not subcommunity_node:
            _add_child(community_node, property_node)


def _load_locations_csv(nodes: dict[str, dict[str, Any]], tier3_set: set[str]) -> None:
    path = Path(LOCATIONS_CSV)
    if not path.exists():
        print(f"[{_timestamp()}] Warning: Locations CSV not found: {LOCATIONS_CSV}")
        return

    with path.open("r", encoding="utf-8", newline="") as file:
        reader = csv.DictReader(file)
        for row in reader:
            canonical = str(row.get("canonical_name") or "").strip()
            if not canonical:
                continue
            level = str(row.get("level") or "").strip().lower()
            parent = str(row.get("parent") or "").strip()

            node = _ensure_node(nodes, canonical, level, tier3_set)
            if not node:
                continue

            if parent and not node.get("parent"):
                parent_node = _ensure_node(nodes, parent, _infer_parent_level(level), tier3_set)
                if parent_node:
                    _add_child(parent_node, node)

            aliases_raw = row.get("aliases") or ""
            for alias in str(aliases_raw).split(","):
                alias_norm = normalize_key(alias)
                if not alias_norm or alias_norm == normalize_key(canonical):
                    continue
                if alias_norm not in node.get("aliases", []):
                    node.setdefault("aliases", []).append(alias_norm)


def _load_coordinates_csv(nodes: dict[str, dict[str, Any]], tier3_set: set[str]) -> None:
    path = Path(COORDINATES_CSV)
    if not path.exists():
        print(f"[{_timestamp()}] Warning: Coordinates CSV not found: {COORDINATES_CSV}")
        return

    with path.open("r", encoding="utf-8", newline="") as file:
        reader = csv.DictReader(file)
        has_level = reader.fieldnames and "level" in reader.fieldnames

        for row in reader:
            canonical = str(row.get("canonical_name") or "").strip()
            if not canonical:
                continue

            level = str(row.get("level") or "").strip().lower() if has_level else ""
            node = _ensure_node(nodes, canonical, level, tier3_set)
            if not node:
                continue

            lat_raw = row.get("lat")
            lng_raw = row.get("lng")
            try:
                lat = float(lat_raw) if lat_raw is not None else None
                lng = float(lng_raw) if lng_raw is not None else None
            except (TypeError, ValueError):
                lat = None
                lng = None

            if lat is not None and lng is not None:
                node["coords"] = (lat, lng)


def _finalize(nodes: dict[str, dict[str, Any]]) -> None:
    for node in nodes.values():
        node["children"] = sorted(set(node.get("children", [])), key=str.lower)
        node["aliases"] = sorted(set(node.get("aliases", [])))


def _build_indexes(nodes: dict[str, dict[str, Any]]) -> tuple[
    dict[str, dict[str, Any]],
    dict[str, list[dict[str, Any]]],
    dict[str, dict[str, Any]],
    dict[str, set[str]],
]:
    tree: dict[str, dict[str, Any]] = {}
    canonical_map: dict[str, dict[str, Any]] = {}
    alias_map: dict[str, list[dict[str, Any]]] = {}
    level_sets = {
        "cities": set(),
        "communities": set(),
        "subcommunities": set(),
        "properties": set(),
    }

    for norm, node in nodes.items():
        tree[node["canonical"]] = node
        canonical_map[norm] = node

        level = (node.get("level") or "").lower()
        if level == "city":
            level_sets["cities"].add(norm)
        elif level == "community":
            level_sets["communities"].add(norm)
        elif level == "subcommunity":
            level_sets["subcommunities"].add(norm)
        elif level == "property":
            level_sets["properties"].add(norm)

        for alias in node.get("aliases", []):
            alias_map.setdefault(alias, []).append(node)

    for alias_nodes in alias_map.values():
        alias_nodes.sort(key=lambda item: item.get("canonical", "").lower())

    return tree, alias_map, canonical_map, level_sets


def _save_cache(
    tree: dict[str, dict[str, Any]],
    alias_map: dict[str, list[dict[str, Any]]],
    level_sets: dict[str, set[str]],
) -> None:
    cache_path = Path(TREE_CACHE)
    cache_path.parent.mkdir(parents=True, exist_ok=True)

    serialized_tree: dict[str, dict[str, Any]] = {}
    for canonical, node in tree.items():
        coords = node.get("coords")
        serialized_tree[canonical] = {
            **node,
            "coords": [coords[0], coords[1]] if coords is not None else None,
        }

    serialized_alias_map = {
        alias: [node.get("canonical") for node in nodes]
        for alias, nodes in alias_map.items()
    }

    payload = {
        "tree": serialized_tree,
        "alias_map": serialized_alias_map,
        "level_sets": {
            key: sorted(list(values)) for key, values in level_sets.items()
        },
    }

    with cache_path.open("w", encoding="utf-8") as file:
        json.dump(payload, file, indent=2)


def _load_cache() -> tuple[
    dict[str, dict[str, Any]],
    dict[str, list[dict[str, Any]]],
    dict[str, dict[str, Any]],
    dict[str, set[str]],
]:
    cache_path = Path(TREE_CACHE)
    with cache_path.open("r", encoding="utf-8") as file:
        payload = json.load(file)

    tree = payload.get("tree", {})
    for node in tree.values():
        coords = node.get("coords")
        if coords is not None:
            node["coords"] = (coords[0], coords[1])

    canonical_map = {
        normalize_key(node.get("canonical", "")): node
        for node in tree.values()
        if node.get("canonical")
    }

    alias_map_serial = payload.get("alias_map", {})
    alias_map: dict[str, list[dict[str, Any]]] = {}
    for alias, canonicals in alias_map_serial.items():
        nodes: list[dict[str, Any]] = []
        for canonical in canonicals:
            node = canonical_map.get(normalize_key(canonical))
            if node:
                nodes.append(node)
        if nodes:
            alias_map[alias] = nodes

    level_sets_serial = payload.get("level_sets", {})
    level_sets = {key: set(values) for key, values in level_sets_serial.items()}

    return tree, alias_map, canonical_map, level_sets


def _cache_is_valid() -> bool:
    cache_path = Path(TREE_CACHE)
    if not cache_path.exists():
        return False

    cache_mtime = cache_path.stat().st_mtime
    for path in (Path(EXCEL_PATH), Path(LOCATIONS_CSV), Path(COORDINATES_CSV)):
        if path.exists() and path.stat().st_mtime > cache_mtime:
            return False

    return True


def _print_counts(level_sets: dict[str, set[str]], source: str) -> None:
    print(source)
    print(f"  cities:        {len(level_sets.get('cities', set()))}")
    print(f"  communities:   {len(level_sets.get('communities', set()))}")
    print(f"  subcommunities:{len(level_sets.get('subcommunities', set()))}")
    print(f"  properties:    {len(level_sets.get('properties', set()))}")


def build_or_load_tree() -> tuple[
    dict[str, dict[str, Any]],
    dict[str, list[dict[str, Any]]],
    dict[str, dict[str, Any]],
    dict[str, set[str]],
]:
    global _TREE, _ALIAS_MAP, _CANONICAL_MAP, _LEVEL_SETS

    if _TREE is not None and _ALIAS_MAP is not None and _CANONICAL_MAP is not None and _LEVEL_SETS is not None:
        return _TREE, _ALIAS_MAP, _CANONICAL_MAP, _LEVEL_SETS

    if not FORCE_REBUILD and _cache_is_valid():
        tree, alias_map, canonical_map, level_sets = _load_cache()
        _TREE, _ALIAS_MAP, _CANONICAL_MAP, _LEVEL_SETS = tree, alias_map, canonical_map, level_sets
        _print_counts(level_sets, "Tree loaded from cache")
        return tree, alias_map, canonical_map, level_sets

    tier3_set = _load_tier3_set()
    nodes: dict[str, dict[str, Any]] = {}

    _load_excel(nodes, tier3_set)
    _load_locations_csv(nodes, tier3_set)
    _load_coordinates_csv(nodes, tier3_set)
    _finalize(nodes)

    tree, alias_map, canonical_map, level_sets = _build_indexes(nodes)
    _save_cache(tree, alias_map, level_sets)

    _TREE, _ALIAS_MAP, _CANONICAL_MAP, _LEVEL_SETS = tree, alias_map, canonical_map, level_sets
    _print_counts(level_sets, "Tree rebuilt from sources")
    return tree, alias_map, canonical_map, level_sets


def get_node(canonical: str) -> dict[str, Any] | None:
    _, _, canonical_map, _ = build_or_load_tree()
    return canonical_map.get(normalize_key(canonical))


def get_children(canonical: str) -> list[str]:
    node = get_node(canonical)
    if not node:
        return []
    return list(node.get("children", []))


def get_parent(canonical: str) -> str | None:
    node = get_node(canonical)
    if not node:
        return None
    return node.get("parent")

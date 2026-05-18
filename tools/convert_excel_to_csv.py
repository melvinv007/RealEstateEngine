"""
convert_excel_to_csv.py
Convert an Excel sheet of locations into locations.csv.
"""

import csv
import os
import sys
from openpyxl import load_workbook

from config import (
    LOCATIONS_CSV,
    LOCATION_CSV_CANONICAL_COLUMN,
    LOCATION_CSV_ALIASES_COLUMN,
    LOCATION_EXCEL_CANONICAL_COLUMN,
    LOCATION_EXCEL_ALIASES_COLUMN,
)


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _split_aliases(value: object) -> list[str]:
    if value is None:
        return []
    raw = value if isinstance(value, str) else str(value)
    return [part.strip() for part in raw.split(",") if part.strip()]


def convert_excel(path: str) -> int:
    workbook = load_workbook(path)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        print("[Converter] No rows found in the Excel sheet.")
        return 1

    header = rows[0]
    header_norm = [_normalize_header(cell) for cell in header]

    canonical_key = _normalize_header(LOCATION_EXCEL_CANONICAL_COLUMN)
    alias_key = _normalize_header(LOCATION_EXCEL_ALIASES_COLUMN)

    if canonical_key not in header_norm:
        print(f"[Converter] Missing canonical column: {LOCATION_EXCEL_CANONICAL_COLUMN}")
        return 1

    canonical_idx = header_norm.index(canonical_key)

    alias_indices = [i for i, value in enumerate(header_norm) if value == alias_key]
    if not alias_indices:
        alias_indices = [i for i, value in enumerate(header_norm) if value.startswith(alias_key) and value]
    if not alias_indices:
        alias_indices = [i for i in range(len(header_norm)) if i != canonical_idx]

    converted: list[tuple[str, str]] = []
    skipped: list[str] = []

    for row_idx, row in enumerate(rows[1:], start=2):
        canonical_cell = row[canonical_idx] if canonical_idx < len(row) else None
        canonical = str(canonical_cell).strip() if canonical_cell is not None else ""
        if not canonical:
            skipped.append(f"Row {row_idx}: missing canonical name")
            continue

        alias_values: list[str] = []
        for idx in alias_indices:
            if idx >= len(row):
                continue
            alias_values.extend(_split_aliases(row[idx]))

        canonical_lower = canonical.strip().lower()
        alias_values = [alias.strip().lower() for alias in alias_values if alias.strip()]
        alias_values = [alias for alias in alias_values if alias != canonical_lower]

        seen = set()
        deduped = []
        for alias in alias_values:
            if alias in seen:
                continue
            seen.add(alias)
            deduped.append(alias)

        converted.append((canonical, ",".join(deduped)))

    output_path = os.path.join(os.path.dirname(__file__), LOCATIONS_CSV)
    with open(output_path, "w", newline="", encoding="utf-8") as file:
        writer = csv.writer(file)
        writer.writerow([LOCATION_CSV_CANONICAL_COLUMN, LOCATION_CSV_ALIASES_COLUMN])
        for canonical, aliases in converted:
            writer.writerow([canonical, aliases])

    print(f"[Converter] Converted {len(converted)} row(s) to {LOCATIONS_CSV}.")
    if skipped:
        print(f"[Converter] Skipped {len(skipped)} row(s):")
        for reason in skipped:
            print(f"  - {reason}")

    return 0


def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: python convert_excel_to_csv.py <locations.xlsx>")
        sys.exit(1)

    sys.exit(convert_excel(sys.argv[1]))


if __name__ == "__main__":
    main()

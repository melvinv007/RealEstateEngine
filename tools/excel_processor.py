"""
excel_processor.py

Reads a location_master.xlsx file and extracts unique location names with
parent relationships, then writes a JSON summary to cache/unique_names.json.

Supports --start-row N (default 2) to skip rows before N (1-based, header is row 1).
"""

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from core.config import LOCATION_MASTER_EXCEL, UNIQUE_NAMES_CACHE

EXCEL_PATH = LOCATION_MASTER_EXCEL
OUTPUT_PATH = UNIQUE_NAMES_CACHE
TIER3_THRESHOLD = 3


def _normalize(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _key(value: Any) -> str:
    return _normalize(value).lower()


def _cell(row: tuple[Any, ...], idx: int) -> str:
    if idx < 0 or idx >= len(row):
        return ""
    return _normalize(row[idx])


def _sort_key(*parts: str) -> tuple[str, ...]:
    return tuple((part or "").lower() for part in parts)


def _load_existing_payload(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {
            "cities": [],
            "communities": [],
            "subcommunities": [],
            "properties": [],
        }
    with path.open("r", encoding="utf-8") as file:
        data = json.load(file)
    return {
        "cities": data.get("cities", []),
        "communities": data.get("communities", []),
        "subcommunities": data.get("subcommunities", []),
        "properties": data.get("properties", []),
    }


def _build_lookup(entries: list[Any]) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for entry in entries:
        if isinstance(entry, str):
            name = _normalize(entry)
        else:
            name = _normalize(entry.get("canonical"))
        if name:
            lookup[name.lower()] = name
    return lookup


def process_excel(excel_path: str = EXCEL_PATH, start_row: int = 2) -> dict[str, Any]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)

    header = next(rows, None)
    if not header:
        raise ValueError("Excel sheet is empty.")

    header_map: dict[str, int] = {}
    for idx, name in enumerate(header):
        key = _key(name)
        if key:
            header_map[key] = idx

    def _col(name: str) -> int:
        key = _key(name)
        if key not in header_map:
            raise ValueError(f"Missing required column: {name}")
        return header_map[key]

    col_city = _col("City")
    col_community = _col("Community")
    col_subcommunity = _col("Subcommunity")
    col_property = _col("Property")

    total_rows = 0

    output_path = Path(OUTPUT_PATH)
    existing_payload = _load_existing_payload(output_path)

    existing_cities = existing_payload.get("cities", [])
    existing_communities = existing_payload.get("communities", [])
    existing_subcommunities = existing_payload.get("subcommunities", [])
    existing_properties = existing_payload.get("properties", [])

    city_lookup = _build_lookup(existing_cities)
    community_lookup = _build_lookup(existing_communities)
    subcommunity_lookup = _build_lookup(existing_subcommunities)
    property_lookup = _build_lookup(existing_properties)

    existing_city_keys = set(city_lookup)
    existing_community_keys = set(community_lookup)
    existing_subcommunity_keys = set(subcommunity_lookup)
    existing_property_keys = set(property_lookup)

    new_city_keys: set[str] = set()

    city_names: dict[str, str] = {}
    community_names: dict[str, str] = {}
    subcommunity_names: dict[str, str] = {}
    property_names: dict[str, str] = {}

    community_entries: dict[tuple[str, str], dict[str, str]] = {}
    subcommunity_entries: dict[tuple[str, str, str], dict[str, str]] = {}
    property_entries: dict[tuple[str, str, str, str], dict[str, str]] = {}
    property_subcommunities: dict[str, set[str]] = {}

    for row_idx, row in enumerate(rows, start=2):
        total_rows += 1
        if row_idx < start_row:
            continue

        city_raw = _cell(row, col_city)
        community_raw = _cell(row, col_community)
        subcommunity_raw = _cell(row, col_subcommunity)
        property_raw = _cell(row, col_property)

        city_key = _key(city_raw)
        community_key = _key(community_raw)
        subcommunity_key = _key(subcommunity_raw)
        property_key = _key(property_raw)

        if city_raw and city_key not in city_lookup:
            city_lookup[city_key] = city_raw
            city_names[city_key] = city_raw
            new_city_keys.add(city_key)

        if community_raw and community_key not in community_lookup:
            community_lookup[community_key] = community_raw
            community_names[community_key] = community_raw

        if subcommunity_raw and subcommunity_key not in subcommunity_lookup:
            subcommunity_lookup[subcommunity_key] = subcommunity_raw
            subcommunity_names[subcommunity_key] = subcommunity_raw

        if property_raw and property_key not in property_lookup:
            property_lookup[property_key] = property_raw
            property_names[property_key] = property_raw

        if community_raw and community_key not in existing_community_keys:
            entry_key = (community_key, city_key)
            if entry_key not in community_entries:
                community_entries[entry_key] = {
                    "canonical": community_lookup.get(community_key, community_raw),
                    "parent_city": city_lookup.get(city_key, city_raw),
                }

        if subcommunity_raw and subcommunity_key not in existing_subcommunity_keys:
            entry_key = (subcommunity_key, community_key, city_key)
            if entry_key not in subcommunity_entries:
                subcommunity_entries[entry_key] = {
                    "canonical": subcommunity_lookup.get(subcommunity_key, subcommunity_raw),
                    "parent_community": community_lookup.get(community_key, community_raw),
                    "parent_city": city_lookup.get(city_key, city_raw),
                }

        if property_raw and property_key not in existing_property_keys:
            entry_key = (property_key, subcommunity_key, community_key, city_key)
            if entry_key not in property_entries:
                property_entries[entry_key] = {
                    "canonical": property_lookup.get(property_key, property_raw),
                    "parent_subcommunity": subcommunity_lookup.get(subcommunity_key, subcommunity_raw),
                    "parent_community": community_lookup.get(community_key, community_raw),
                    "parent_city": city_lookup.get(city_key, city_raw),
                }

            # Only count non-empty subcommunities toward tier3 detection.
            if subcommunity_key:
                property_subcommunities.setdefault(property_key, set()).add(subcommunity_key)
    cities_new = sorted((city_lookup[key] for key in new_city_keys), key=str.lower)

    communities_new = sorted(
        community_entries.values(),
        key=lambda entry: _sort_key(entry.get("canonical", ""), entry.get("parent_city", "")),
    )

    subcommunities_new = sorted(
        subcommunity_entries.values(),
        key=lambda entry: _sort_key(
            entry.get("canonical", ""),
            entry.get("parent_community", ""),
            entry.get("parent_city", ""),
        ),
    )

    properties_new: list[dict[str, Any]] = []
    for (property_key, _, _, _), entry in property_entries.items():
        subcommunity_count = len(property_subcommunities.get(property_key, set()))
        properties_new.append({
            **entry,
            "tier3": subcommunity_count >= TIER3_THRESHOLD,
            "subcommunity_count": subcommunity_count,
        })

    properties_new.sort(
        key=lambda entry: _sort_key(
            entry.get("canonical", ""),
            entry.get("parent_subcommunity", ""),
            entry.get("parent_community", ""),
            entry.get("parent_city", ""),
        )
    )

    payload = {
        "cities": existing_cities + cities_new,
        "communities": existing_communities + communities_new,
        "subcommunities": existing_subcommunities + subcommunities_new,
        "properties": existing_properties + properties_new,
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as file:
        json.dump(payload, file, indent=2)

    tier3_count = sum(1 for entry in payload["properties"] if entry.get("tier3"))

    print(f"Total rows in Excel: {total_rows}")
    print(f"Unique cities count: {len(payload['cities'])}")
    print(f"Unique communities count: {len(payload['communities'])}")
    print(f"Unique subcommunities count: {len(payload['subcommunities'])}")
    print(f"Unique properties count: {len(payload['properties'])}")
    print(f"Tier 3 properties count: {tier3_count}")

    return payload


def main() -> int:
    parser = argparse.ArgumentParser(description="Extract unique location names from Excel")
    parser.add_argument("--start-row", type=int, default=2, help="Start row (1-based, header is row 1)")
    parser.add_argument("--excel-path", type=str, default=EXCEL_PATH, help="Path to location_master.xlsx")
    args = parser.parse_args()

    try:
        process_excel(excel_path=args.excel_path, start_row=args.start_row)
        return 0
    except Exception as exc:
        print(f"[Error] {exc}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())

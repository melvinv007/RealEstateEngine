"""
geocode_locations.py

Reads an Excel file with columns: City | Community | Subcommunity | Property
(Property may be empty). For each row, builds a canonical name and geocodes it
using Nominatim. Adds new entries to data/coordinates.csv — skips if already present.
Excel file is never modified.

Usage:
    python tools/geocode_locations.py <path_to_excel.xlsx> --start-row <N> (optional, default 2 to skip header)
"""

import csv
import json
import os
import sys
import time
from pathlib import Path

from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut, GeocoderServiceError
from openpyxl import load_workbook

sys.path.append(str(Path(__file__).resolve().parent.parent))

from core.config import COORDINATES_CSV, GEOCODE_FAILURES_FILE

csv.field_size_limit(sys.maxsize)

# ── Configuration ──────────────────────────────────────────────────────────────

_PROJECT_ROOT = Path(__file__).resolve().parent.parent

# Excel column headers (case-insensitive)
COL_CITY          = "City"
COL_COMMUNITY     = "Community"
COL_SUBCOMMUNITY  = "Subcommunity"
COL_PROPERTY      = "Property"

# What becomes the canonical_name key in coordinates.csv
# Options: "community" | "subcommunity" | "property"
# - "community"    → Al Aweer, Zabeel, Reem
# - "subcommunity" → Al Aweer 1, Al Murooj Complex, Mira
# - "property"     → Desert Palm, Gardenia (falls back to subcommunity if empty)
CANONICAL_LEVEL = "subcommunity"

# Which columns to include in the geocode query (always joined with City)
# Options: "community" | "subcommunity" | "property"
# "subcommunity" → "Al Aweer 1, Al Aweer, Dubai, UAE"
# "property"     → "Desert Palm, Al Aweer 1, Al Aweer, Dubai, UAE" (if property exists)
GEOCODE_LEVEL = "subcommunity"

# Appended to every geocode query; set "" to disable
APPEND_SUFFIX = "UAE"

# Skip rows where property is empty (useful if you only want named buildings)
SKIP_IF_NO_PROPERTY = False

# Skip rows where subcommunity is empty
SKIP_IF_NO_SUBCOMMUNITY = True

# Deduplicate: if two rows produce the same canonical_name, geocode only the first
SKIP_DUPLICATE_CANONICALS = True

# data/coordinates.csv settings
COORDINATES_CSV_PATH = COORDINATES_CSV
CSV_CANONICAL_COLUMN = "canonical_name"
CSV_LEVEL_COLUMN     = "level"
CSV_LAT_COLUMN       = "lat"
CSV_LNG_COLUMN       = "lng"

UNIQUE_NAMES_CACHE = "cache/unique_names.json"

# Nominatim settings
NOMINATIM_DELAY_SECONDS = 2.0
NOMINATIM_USER_AGENT    = "dubai_realestate_geocoder_v1"

# Misc
VERBOSE = True      # Print a status line for every row
DRY_RUN = False     # True = print results but write nothing

TIER3_SKIP = True  # If True, skip geocoding properties marked as tier3 in unique_names.json
# ──────────────────────────────────────────────────────────────────────────────


_geolocator = Nominatim(user_agent=NOMINATIM_USER_AGENT, timeout=10)
_last_call: float = 0.0


# ── Nominatim ─────────────────────────────────────────────────────────────────

def _rate_limit() -> None:
    global _last_call
    elapsed = time.time() - _last_call
    if elapsed < NOMINATIM_DELAY_SECONDS:
        time.sleep(NOMINATIM_DELAY_SECONDS - elapsed)
    _last_call = time.time()


def _geocode(query: str) -> tuple[float, float] | None:
    _rate_limit()
    try:
        result = _geolocator.geocode(query)
        return (result.latitude, result.longitude) if result else None
    except GeocoderTimedOut:
        if VERBOSE:
            print(f"\n  [Geocoder] Timeout — retrying '{query}'...")
        time.sleep(NOMINATIM_DELAY_SECONDS * 2)
        try:
            result = _geolocator.geocode(query)
            return (result.latitude, result.longitude) if result else None
        except Exception:
            return None
    except (GeocoderServiceError, Exception) as e:
        if VERBOSE:
            print(f"\n  [Geocoder] Error for '{query}': {e}")
        return None


# ── Query + canonical builders ─────────────────────────────────────────────────

def _build_query(city: str, community: str, subcommunity: str, property_: str) -> str:
    """Build the geocode search string from available fields."""
    parts = []

    if GEOCODE_LEVEL == "property" and property_:
        parts.append(property_)

    if GEOCODE_LEVEL in ("property", "subcommunity") and subcommunity:
        parts.append(subcommunity)

    if community:
        parts.append(community)

    if city:
        parts.append(city)

    if APPEND_SUFFIX:
        parts.append(APPEND_SUFFIX)

    return ", ".join(parts)


def _build_canonical(community: str, subcommunity: str, property_: str) -> str | None:
    """Build the canonical_name that will be stored in coordinates.csv."""
    if CANONICAL_LEVEL == "property":
        return property_ if property_ else (subcommunity or community or None)
    elif CANONICAL_LEVEL == "subcommunity":
        return subcommunity if subcommunity else (community or None)
    elif CANONICAL_LEVEL == "community":
        return community or None
    return None


# ── CSV helpers ────────────────────────────────────────────────────────────────

def _ensure_level_column(csv_path: str) -> None:
    if not os.path.exists(csv_path) or os.path.getsize(csv_path) == 0:
        return

    with open(csv_path, "r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        header = reader.fieldnames or []
        if CSV_LEVEL_COLUMN in header:
            return
        rows = list(reader)

    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([CSV_CANONICAL_COLUMN, CSV_LEVEL_COLUMN, CSV_LAT_COLUMN, CSV_LNG_COLUMN])
        for row in rows:
            writer.writerow([
                row.get(CSV_CANONICAL_COLUMN, ""),
                "unknown",
                row.get(CSV_LAT_COLUMN, ""),
                row.get(CSV_LNG_COLUMN, ""),
            ])


def _load_csv_canonicals(csv_path: str) -> set[str]:
    """Return set of lower-cased canonical names already in coordinates.csv."""
    existing: set[str] = set()
    if not os.path.exists(csv_path):
        return existing
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = row.get(CSV_CANONICAL_COLUMN, "").strip().lower()
            if name:
                existing.add(name)
    return existing


def _append_to_csv(csv_path: str, canonical: str, level: str, lat: float, lng: float) -> None:
    """Append one row to coordinates.csv, writing the header if file is new."""
    file_exists = os.path.exists(csv_path)
    with open(csv_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow([CSV_CANONICAL_COLUMN, CSV_LEVEL_COLUMN, CSV_LAT_COLUMN, CSV_LNG_COLUMN])
        writer.writerow([canonical, level, lat, lng])


# ── Helpers ────────────────────────────────────────────────────────────────────

def _normalize_header(v: object) -> str:
    return "" if v is None else str(v).strip().lower()


def _cell_str(row: tuple, idx: int) -> str:
    """Safely extract a string cell value from a row tuple."""
    if idx is None or idx >= len(row):
        return ""
    v = row[idx]
    return str(v).strip() if v is not None else ""


def _load_tier3_properties(cache_path: str) -> set[str]:
    if not os.path.exists(cache_path):
        return set()

    with open(cache_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    tier3: set[str] = set()
    for entry in data.get("properties", []):
        if entry.get("tier3"):
            name = str(entry.get("canonical") or "").strip().lower()
            if name:
                tier3.add(name)
    return tier3


def _load_failures(cache_path: str) -> set[str]:
    if not os.path.exists(cache_path):
        return set()
    try:
        with open(cache_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return set()

    failures: set[str] = set()
    if isinstance(data, list):
        for entry in data:
            name = str(entry or "").strip().lower()
            if name:
                failures.add(name)
    elif isinstance(data, dict):
        for entry in data.get("failures", []):
            name = str(entry or "").strip().lower()
            if name:
                failures.add(name)
    return failures


def _save_failures(cache_path: str, failures: set[str]) -> None:
    Path(cache_path).parent.mkdir(parents=True, exist_ok=True)
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(sorted(failures), f, indent=2)


def _build_query_parts(parts: list[str]) -> str:
    return ", ".join([part for part in parts if part])


# ── Main ───────────────────────────────────────────────────────────────────────

def old_geocode_excel(excel_path: str) -> int:
    if not os.path.exists(excel_path):
        print(f"[Error] File not found: {excel_path}")
        return 1

    wb = load_workbook(excel_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        print("[Error] Excel sheet is empty.")
        return 1

    # Resolve column indices from header row
    header_norm = [_normalize_header(cell) for cell in rows[0]]

    def _col(name: str) -> int | None:
        key = _normalize_header(name)
        return header_norm.index(key) if key in header_norm else None

    col_city         = _col(COL_CITY)
    col_community    = _col(COL_COMMUNITY)
    col_subcommunity = _col(COL_SUBCOMMUNITY)
    col_property     = _col(COL_PROPERTY)

    missing = [
        name for name, idx in [
            (COL_CITY, col_city),
            (COL_COMMUNITY, col_community),
            (COL_SUBCOMMUNITY, col_subcommunity),
        ] if idx is None
    ]
    if missing:
        print(f"[Error] Missing required columns: {missing}")
        print(f"  Found headers: {[c for c in rows[0] if c]}")
        return 1

    if col_property is None:
        print(f"[Warning] Column '{COL_PROPERTY}' not found — property treated as empty for all rows.")

    _ensure_level_column(COORDINATES_CSV_PATH)

    # Load what's already in coordinates.csv
    csv_existing = _load_csv_canonicals(COORDINATES_CSV_PATH)
    print(f"[Setup] {len(csv_existing)} entries already in {COORDINATES_CSV_PATH}")

    cache_path = str(_PROJECT_ROOT / UNIQUE_NAMES_CACHE)
    cache_exists = os.path.exists(cache_path)
    tier3_properties = _load_tier3_properties(cache_path)
    if not cache_exists and VERBOSE:
        print("[Warning] Tier3 cache not found; property filtering disabled.")

    failures = _load_failures(GEOCODE_FAILURES_FILE)

    stats = {
        "geocoded": {"city": 0, "community": 0, "subcommunity": 0, "property": 0},
        "failed": {"city": 0, "community": 0, "subcommunity": 0, "property": 0},
        "skipped": {"city": 0, "community": 0, "subcommunity": 0, "property": 0},
    }

    def _bump(bucket: str, level: str) -> None:
        stats[bucket][level] = stats[bucket].get(level, 0) + 1

    def _process(level: str, canonical: str, query: str, row_idx: int) -> None:
        canonical = canonical.strip()
        if not canonical:
            return

        canonical_lower = canonical.lower()

        if canonical_lower in csv_existing:
            _bump("skipped", level)
            if VERBOSE:
                print(f"  Row {row_idx:>4}: SKIP  ({level} already in CSV) — {canonical}")
            return

        if canonical_lower in failures:
            _bump("skipped", level)
            if VERBOSE:
                print(f"  Row {row_idx:>4}: SKIP  ({level} in failures cache) — {canonical}")
            return

        if VERBOSE:
            print(f"  Row {row_idx:>4}: QUERY [{level}] '{query}' ...", end=" ", flush=True)

        coords = _geocode(query)
        if coords:
            lat, lng = coords
            if VERBOSE:
                print(f"→ ({lat:.5f}, {lng:.5f})")
            if not DRY_RUN:
                _append_to_csv(COORDINATES_CSV_PATH, canonical, level, lat, lng)
            csv_existing.add(canonical_lower)
            if canonical_lower in failures:
                failures.remove(canonical_lower)
                _save_failures(GEOCODE_FAILURES_FILE, failures)
            _bump("geocoded", level)
        else:
            if VERBOSE:
                print("→ FAILED")
            _bump("failed", level)
            failures.add(canonical_lower)
            _save_failures(GEOCODE_FAILURES_FILE, failures)
            print(f"  [Geocoder] FAILED: {canonical} — added to failures cache")

    for row_idx, row in enumerate(rows[1:], start=2):
        city         = _cell_str(row, col_city)
        community    = _cell_str(row, col_community)
        subcommunity = _cell_str(row, col_subcommunity)
        property_    = _cell_str(row, col_property) if col_property is not None else ""

        if city:
            query = _build_query_parts([city, APPEND_SUFFIX])
            _process("city", city, query, row_idx)

        if community:
            query = _build_query_parts([community, city, APPEND_SUFFIX])
            _process("community", community, query, row_idx)

        if subcommunity:
            query = _build_query_parts([subcommunity, community, city, APPEND_SUFFIX])
            _process("subcommunity", subcommunity, query, row_idx)

        if property_:
            is_tier3 = property_.strip().lower() in tier3_properties
            if TIER3_SKIP and is_tier3:
                if VERBOSE:
                    print(f"  Row {row_idx:>4}: SKIP  (property tier3) — {property_}")
            else:
                query = _build_query_parts([property_, subcommunity, community, city, APPEND_SUFFIX])
                _process("property", property_, query, row_idx)

    print("\n[Done]")
    for label, bucket in (
        ("Geocoded", "geocoded"),
        ("Failed", "failed"),
        ("Skipped (already existed)", "skipped"),
    ):
        print(f"  {label}:")
        for level in ("city", "community", "subcommunity", "property"):
            print(f"    {level:12s}: {stats[bucket].get(level, 0)}")
    if DRY_RUN:
        print("  [Dry Run] Nothing written.")

    return 0

def geocode_excel(excel_path: str, start_row: int = 2) -> int:
    """Geocode rows with index >= start_row (1-based, header is row 1)."""
    if not os.path.exists(excel_path):
        print(f"[Error] File not found: {excel_path}")
        return 1

    wb = load_workbook(excel_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        print("[Error] Excel sheet is empty.")
        return 1

    header_norm = [_normalize_header(cell) for cell in rows[0]]

    def _col(name: str) -> int | None:
        key = _normalize_header(name)
        return header_norm.index(key) if key in header_norm else None

    col_city         = _col(COL_CITY)
    col_community    = _col(COL_COMMUNITY)
    col_subcommunity = _col(COL_SUBCOMMUNITY)
    col_property     = _col(COL_PROPERTY)

    missing = [
        name for name, idx in [
            (COL_CITY, col_city),
            (COL_COMMUNITY, col_community),
            (COL_SUBCOMMUNITY, col_subcommunity),
        ] if idx is None
    ]
    if missing:
        print(f"[Error] Missing required columns: {missing}")
        print(f"  Found headers: {[c for c in rows[0] if c]}")
        return 1

    if col_property is None:
        print(f"[Warning] Column '{COL_PROPERTY}' not found — property treated as empty for all rows.")

    _ensure_level_column(COORDINATES_CSV_PATH)

    csv_existing = _load_csv_canonicals(COORDINATES_CSV_PATH)
    print(f"[Setup] {len(csv_existing)} entries already in {COORDINATES_CSV_PATH}")

    cache_path = str(_PROJECT_ROOT / UNIQUE_NAMES_CACHE)
    cache_exists = os.path.exists(cache_path)
    tier3_properties = _load_tier3_properties(cache_path)
    if not cache_exists and VERBOSE:
        print("[Warning] Tier3 cache not found; property filtering disabled.")

    failures = _load_failures(GEOCODE_FAILURES_FILE)

    if start_row > 2:
        print(f"[Setup] Skipping rows 2–{start_row - 1} (--start-row {start_row})")

    stats = {
        "geocoded": {"city": 0, "community": 0, "subcommunity": 0, "property": 0},
        "failed":   {"city": 0, "community": 0, "subcommunity": 0, "property": 0},
        "skipped":  {"city": 0, "community": 0, "subcommunity": 0, "property": 0},
    }

    def _bump(bucket: str, level: str) -> None:
        stats[bucket][level] = stats[bucket].get(level, 0) + 1

    def _process(level: str, canonical: str, query: str, row_idx: int) -> None:
        canonical = canonical.strip()
        if not canonical:
            return

        canonical_lower = canonical.lower()

        if canonical_lower in csv_existing:
            _bump("skipped", level)
            if VERBOSE:
                print(f"  Row {row_idx:>4}: SKIP  ({level} already in CSV) — {canonical}")
            return

        if canonical_lower in failures:
            _bump("skipped", level)
            if VERBOSE:
                print(f"  Row {row_idx:>4}: SKIP  ({level} in failures cache) — {canonical}")
            return

        if VERBOSE:
            print(f"  Row {row_idx:>4}: QUERY [{level}] '{query}' ...", end=" ", flush=True)

        coords = _geocode(query)
        if coords:
            lat, lng = coords
            if VERBOSE:
                print(f"→ ({lat:.5f}, {lng:.5f})")
            if not DRY_RUN:
                _append_to_csv(COORDINATES_CSV_PATH, canonical, level, lat, lng)
            csv_existing.add(canonical_lower)
            if canonical_lower in failures:
                failures.remove(canonical_lower)
                _save_failures(GEOCODE_FAILURES_FILE, failures)
            _bump("geocoded", level)
        else:
            if VERBOSE:
                print("→ FAILED")
            _bump("failed", level)
            failures.add(canonical_lower)
            _save_failures(GEOCODE_FAILURES_FILE, failures)
            print(f"  [Geocoder] FAILED: {canonical} — added to failures cache")

    for row_idx, row in enumerate(rows[1:], start=2):
        if row_idx < start_row:
            continue

        city         = _cell_str(row, col_city)
        community    = _cell_str(row, col_community)
        subcommunity = _cell_str(row, col_subcommunity)
        property_    = _cell_str(row, col_property) if col_property is not None else ""

        if city:
            query = _build_query_parts([city, APPEND_SUFFIX])
            _process("city", city, query, row_idx)

        if community:
            query = _build_query_parts([community, city, APPEND_SUFFIX])
            _process("community", community, query, row_idx)

        if subcommunity:
            query = _build_query_parts([subcommunity, community, city, APPEND_SUFFIX])
            _process("subcommunity", subcommunity, query, row_idx)

        if property_:
            is_tier3 = property_.strip().lower() in tier3_properties
            if TIER3_SKIP and is_tier3:
                if VERBOSE:
                    print(f"  Row {row_idx:>4}: SKIP  (property tier3) — {property_}")
            else:
                query = _build_query_parts([property_, subcommunity, community, city, APPEND_SUFFIX])
                _process("property", property_, query, row_idx)

    print("\n[Done]")
    for label, bucket in (
        ("Geocoded", "geocoded"),
        ("Failed", "failed"),
        ("Skipped (already existed)", "skipped"),
    ):
        print(f"  {label}:")
        for level in ("city", "community", "subcommunity", "property"):
            print(f"    {level:12s}: {stats[bucket].get(level, 0)}")
    if DRY_RUN:
        print("  [Dry Run] Nothing written.")

    return 0

def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: python tools/geocode_locations.py <path_to_locations.xlsx> --start-row <N> (default N=2)")
        sys.exit(1)
    import argparse
    parser = argparse.ArgumentParser(
        description="Geocode locations from Excel into coordinates.csv"
    )
    parser.add_argument("excel", help="Path to Excel file")
    parser.add_argument(
        "--start-row",
        type=int,
        default=2,
        help="Excel row number to start from (default: 2, i.e. first data row)"
    )
    args = parser.parse_args()
    sys.exit(geocode_excel(args.excel, start_row=args.start_row))


if __name__ == "__main__":
    main()
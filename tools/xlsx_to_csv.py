"""
xlsx_to_csv.py
A robust utility script to convert Excel (.xlsx) files into CSV format.
Features:
---------
- Configurable input/output paths
- Customizable sheet selection
- Configurable start row
- Optional header handling
- Optional trimming of whitespace
- Graceful Ctrl+C interruption handling
- Detailed logging and progress messages

Requirements: pip install openpyxl

Usage:
------
1. Edit the CONFIG section below.
2. Run:
    python xlsx_to_csv.py

"""

from __future__ import annotations

import csv
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


# =============================================================================
# CONFIGURATION
# =============================================================================

# Input Excel file (.xlsx)
INPUT_XLSX_PATH = "input.xlsx"

# Output CSV file
OUTPUT_CSV_PATH = "data/project_master.csv"

# Sheet selection:
# - Use sheet name: "Sheet1"
# - OR use sheet index: 0 (first sheet), 1, 2, etc.
SHEET_SELECTION: str | int = 0

# Row number to START reading from (1-based indexing)
START_ROW = 1

# Maximum rows to export.
# Set to None to export all rows.
MAX_ROWS: int | None = None

# Include completely empty rows in CSV
INCLUDE_EMPTY_ROWS = False

# Remove leading/trailing whitespace from cell values
TRIM_WHITESPACE = True

# Convert None values to empty string
CONVERT_NONE_TO_EMPTY = True

# Output CSV encoding
CSV_ENCODING = "utf-8"

# CSV delimiter
CSV_DELIMITER = ","

# CSV quote handling
CSV_QUOTECHAR = '"'

# Line terminator
CSV_LINETERMINATOR = "\n"

# Overwrite existing output file
OVERWRITE_OUTPUT = True

# Print progress every N rows
PROGRESS_EVERY_N_ROWS = 1000

# =============================================================================
# END CONFIGURATION
# =============================================================================


def print_header() -> None:
    """Prints a startup banner."""
    print("=" * 70)
    print("XLSX TO CSV CONVERTER")
    print("=" * 70)


def validate_paths(input_path: Path, output_path: Path) -> None:
    """
    Validates input and output paths.

    Args:
        input_path: Path to input XLSX file.
        output_path: Path to output CSV file.

    Raises:
        FileNotFoundError: If input file does not exist.
        ValueError: If file extensions are invalid.
        FileExistsError: If output exists and overwrite disabled.
    """

    if not input_path.exists():
        raise FileNotFoundError(f"Input file does not exist: {input_path}")

    if not input_path.is_file():
        raise ValueError(f"Input path is not a file: {input_path}")

    if input_path.suffix.lower() != ".xlsx":
        raise ValueError(
            f"Input file must be a .xlsx file, got: {input_path.suffix}"
        )

    if output_path.suffix.lower() != ".csv":
        raise ValueError(
            f"Output file must be a .csv file, got: {output_path.suffix}"
        )

    if output_path.exists() and not OVERWRITE_OUTPUT:
        raise FileExistsError(
            f"Output file already exists and overwrite is disabled: "
            f"{output_path}"
        )


def get_worksheet(workbook, selection: str | int):
    """
    Returns worksheet based on sheet name or index.

    Args:
        workbook: OpenPyXL workbook object.
        selection: Sheet name or sheet index.

    Returns:
        Worksheet object.

    Raises:
        ValueError: If sheet does not exist.
    """

    if isinstance(selection, int):
        sheet_names = workbook.sheetnames

        if selection < 0 or selection >= len(sheet_names):
            raise ValueError(
                f"Sheet index out of range. "
                f"Workbook contains {len(sheet_names)} sheets."
            )

        return workbook[sheet_names[selection]]

    if isinstance(selection, str):
        if selection not in workbook.sheetnames:
            raise ValueError(
                f"Sheet '{selection}' not found. "
                f"Available sheets: {workbook.sheetnames}"
            )

        return workbook[selection]

    raise ValueError("SHEET_SELECTION must be a string or integer.")


def clean_cell_value(value: Any) -> str:
    """
    Cleans and converts cell value to string.

    Args:
        value: Raw cell value.

    Returns:
        Cleaned string representation.
    """

    if value is None:
        return "" if CONVERT_NONE_TO_EMPTY else "None"

    value_str = str(value)

    if TRIM_WHITESPACE:
        value_str = value_str.strip()

    return value_str


def is_row_empty(row: list[str]) -> bool:
    """
    Checks whether an entire row is empty.

    Args:
        row: List of row values.

    Returns:
        True if all cells are empty, False otherwise.
    """

    return all(cell == "" for cell in row)


def ensure_output_directory(output_path: Path) -> None:
    """
    Creates output directory if it does not exist.

    Args:
        output_path: CSV output path.
    """

    output_path.parent.mkdir(parents=True, exist_ok=True)


def convert_xlsx_to_csv() -> int:
    """
    Main conversion logic.

    Returns:
        Exit code:
        0 = Success
        1 = Failure
    """

    input_path = Path(INPUT_XLSX_PATH)
    output_path = Path(OUTPUT_CSV_PATH)

    print_header()

    print(f"Input XLSX : {input_path}")
    print(f"Output CSV : {output_path}")
    print(f"Sheet       : {SHEET_SELECTION}")
    print(f"Start Row   : {START_ROW}")
    print()

    validate_paths(input_path, output_path)
    ensure_output_directory(output_path)

    print("Opening workbook...")

    workbook = load_workbook(
        filename=input_path,
        read_only=True,
        data_only=True,
    )

    worksheet = get_worksheet(workbook, SHEET_SELECTION)

    print(f"Using worksheet: {worksheet.title}")

    max_row = worksheet.max_row or 0
    max_column = worksheet.max_column or 0

    print(f"Detected rows    : {max_row}")
    print(f"Detected columns : {max_column}")

    if max_row == 0 or max_column == 0:
        print("Excel sheet is empty. Nothing to export.")
        return 0

    if START_ROW < 1:
        raise ValueError("START_ROW must be >= 1.")

    if START_ROW > max_row:
        raise ValueError(
            f"START_ROW ({START_ROW}) exceeds worksheet row count ({max_row})."
        )

    rows_written = 0
    rows_processed = 0

    print()
    print("Starting conversion...")
    print()

    with open(
        output_path,
        mode="w",
        encoding=CSV_ENCODING,
        newline="",
    ) as csv_file:

        writer = csv.writer(
            csv_file,
            delimiter=CSV_DELIMITER,
            quotechar=CSV_QUOTECHAR,
            quoting=csv.QUOTE_MINIMAL,
            lineterminator=CSV_LINETERMINATOR,
        )

        for excel_row_index, row in enumerate(
            worksheet.iter_rows(
                min_row=START_ROW,
                values_only=True,
            ),
            start=START_ROW,
        ):

            rows_processed += 1

            cleaned_row = [
                clean_cell_value(cell)
                for cell in row
            ]

            if not INCLUDE_EMPTY_ROWS and is_row_empty(cleaned_row):
                continue

            writer.writerow(cleaned_row)
            rows_written += 1

            if (
                PROGRESS_EVERY_N_ROWS > 0
                and rows_processed % PROGRESS_EVERY_N_ROWS == 0
            ):
                print(
                    f"Processed {rows_processed:,} rows | "
                    f"Written {rows_written:,} rows..."
                )

            if MAX_ROWS is not None and rows_written >= MAX_ROWS:
                print()
                print(f"Reached MAX_ROWS limit ({MAX_ROWS}).")
                break

    workbook.close()

    print()
    print("=" * 70)
    print("CONVERSION COMPLETED SUCCESSFULLY")
    print("=" * 70)
    print(f"Rows processed : {rows_processed:,}")
    print(f"Rows written   : {rows_written:,}")
    print(f"Output saved   : {output_path.resolve()}")

    return 0


def main() -> int:
    """
    Program entry point.

    Handles:
    - Ctrl+C interruption
    - File errors
    - Unexpected runtime errors

    Returns:
        Process exit code.
    """

    try:
        return convert_xlsx_to_csv()

    except KeyboardInterrupt:
        print()
        print()
        print("Operation interrupted by user (Ctrl+C).")
        return 1

    except FileNotFoundError as exc:
        print()
        print(f"File error: {exc}")
        return 1

    except PermissionError as exc:
        print()
        print(f"Permission error: {exc}")
        print("Make sure the file is not open in Excel.")
        return 1

    except ValueError as exc:
        print()
        print(f"Validation error: {exc}")
        return 1

    except Exception as exc:
        print()
        print("Unexpected error occurred.")
        print(f"Error type: {type(exc).__name__}")
        print(f"Error message: {exc}")
        return 1


if __name__ == "__main__":
    sys.exit(main())